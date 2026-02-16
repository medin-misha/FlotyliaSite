from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload
from sqlalchemy import select, func, or_
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import io
from core.models import User, Transport, Contract, File, Document
from services.error_handlers import DBErrorHandler
from fastapi import HTTPException
from sqlalchemy import Result


async def _get_model_data(session: AsyncSession, model) -> list:
    try:
        stmt = select(model)
        result: Result = await session.execute(stmt)
        items = result.scalars().all()
        
        # Convert SQLAlchemy objects to dicts
        data = []
        for item in items:
            item_dict = {}
            for column in item.__table__.columns:
                value = getattr(item, column.name)
                # Convert date/time to string for Excel compatibility if needed, 
                # but pandas usually handles it well. 
                item_dict[column.name] = str(value) if value is not None else None
            data.append(item_dict)
        return data
    except Exception as err:
        DBErrorHandler.handle(err=err, model=model)


async def export_to_exel(session: AsyncSession) -> StreamingResponse:
    models_to_export = {
        "Users": User,
        "Contracts": Contract,
        "Transports": Transport,
        "Files": File,
        "Documents": Document,
    }

    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, model in models_to_export.items():
            data = await _get_model_data(session, model)
            if not data:
                # Create an empty dataframe with column names if no data
                columns = [c.name for c in model.__table__.columns]
                df = pd.DataFrame(columns=columns)
            else:
                df = pd.DataFrame(data)
                
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            # Formatting
            sheet = writer.sheets[sheet_name]
            fill = PatternFill(start_color="4B85E7", end_color="4B85E7", fill_type="solid")
            font = Font(color="FFFFFF", bold=True)
            
            # Set column widths
            for i in range(1, sheet.max_column + 1):
                symbol = get_column_letter(i)
                sheet.column_dimensions[symbol].width = 20
                
            # Header styling
            for cell in sheet[1]:
                cell.font = font
                cell.fill = fill

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=database_export.xlsx"},
    )

